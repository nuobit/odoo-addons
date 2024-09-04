# Copyright NuoBiT Solutions, S.L. (<https://www.nuobit.com>)
# Eric Antones <eantones@nuobit.com>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl)

import base64
from io import BytesIO

import xlrd
from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError

MAP_TYPES = {
    xlrd.XL_CELL_TEXT: _("Text"),
    xlrd.XL_CELL_NUMBER: _("Numeric"),
    xlrd.XL_CELL_DATE: _("Date"),
    xlrd.XL_CELL_BOOLEAN: _("Boolean"),
    xlrd.XL_CELL_ERROR: _("Error"),
    xlrd.XL_CELL_BLANK: _("Blank"),
    xlrd.XL_CELL_EMPTY: _("Empty"),
}


def _get_cell(sheet, row_index, col_index, ctype, nullable=False):
    value = sheet.cell_value(row_index, col_index)
    cellname = xlrd.cellname(row_index, col_index)
    current_ctype = sheet.cell_type(row_index, col_index)
    if ctype == current_ctype == xlrd.XL_CELL_TEXT:
        value = value.strip() or None
        if not nullable and not value:
            raise UserError(_("Null value in cell %s") % cellname)
    elif ctype == current_ctype == xlrd.XL_CELL_NUMBER:
        pass
    else:
        if current_ctype in MAP_TYPES:
            current_ctype = MAP_TYPES[current_ctype]
        if ctype in MAP_TYPES:
            ctype = MAP_TYPES[ctype]
        raise UserError(
            _("Wrong format in cell %s. Expected %s, found %s")
            % (cellname, ctype, current_ctype)
        )

    return value


class MsgLog:
    def __init__(self, mtypes):
        self.content = {x: [] for x in mtypes.keys()}
        self.mtypes = mtypes

    def add_msg(self, mtype, msg):
        self.content[mtype].append(msg)

    def get_type_desc(self, mtype):
        return self.mtypes[mtype]

    def get_msgs(self):
        return {k: v for k, v in self.content.items() if v}

    @property
    def has_msgs(self):
        return bool(self.get_msgs())


class StockPickingImportSerials(models.TransientModel):
    _name = "stock.picking.import.serials"
    _description = "Stock Picking Import Serials"

    datas = fields.Binary(
        string="File",
        attachment=True,
        required=True,
        help="The fields you can import are:\n -Default Code\n -Serial Number",
    )

    datas_fname = fields.Char(string="Filename", required=True)

    result = fields.Text(string="Result", readonly=True)

    state = fields.Selection(
        [
            ("params", _("Parameters")),
            ("result", _("Result")),
        ],
        string="Status",
        default="params",
        readonly=True,
        required=True,
        copy=False,
    )

    def _get_product_serial_xls(self, sheet):
        max_columns = sheet.ncols
        LAST_COLUMN = 2
        if max_columns < LAST_COLUMN:
            raise UserError(
                _("Incorrect format file, the number of columns must be 2 minimum")
            )
        data = []
        for row_index in range(max_columns):
            row_values = sheet.row_values(row_index, 0, LAST_COLUMN)
            if row_index == 0:
                types = set(sheet.row_types(row_index, 0, LAST_COLUMN))
                if types != {xlrd.XL_CELL_TEXT}:
                    raise UserError(_("All fields of the header row must be text"))
                else:
                    if not all(
                        map(lambda x: x.strip() or None is not None, row_values)
                    ):
                        raise UserError(
                            _("The fields on the header row must not be null")
                        )
            else:
                data.append(
                    tuple(
                        [
                            _get_cell(sheet, row_index, x, xlrd.XL_CELL_TEXT)
                            for x in range(max_columns)
                        ]
                    )
                )
            return data

    def _get_product_serial_xlsx(self, sheet):
        max_columns = sheet.max_column
        LAST_COLUMN = 2
        if max_columns < LAST_COLUMN:
            raise UserError(
                _("Incorrect format file, the number of columns must be 2 minimum")
            )
        data = []
        first_row = True
        for row in sheet.rows:
            if first_row:
                types = {c.data_type for c in row}
                if types != {"s"}:
                    raise UserError(_("All fields of the header row must be text"))
                if any(not c.internal_value for c in row):
                    raise UserError(_("The fields on the header row must not be null"))
                first_row = False
                continue
            data.append(tuple([c.internal_value for c in row]))
        return data

    def _prepare_additional_tracking_values(self, data, company):
        return {}

    def _prepare_data(self, data, company):
        unique_tracking_numbers = {}
        product_serial = {}
        for row in data:
            default_code = row[0]
            tracking_number = row[1]
            # check serial uniqueness by product
            if default_code not in unique_tracking_numbers:
                unique_tracking_numbers[default_code] = {tracking_number}
            else:
                if tracking_number in unique_tracking_numbers[default_code]:
                    raise ValidationError(
                        _("The tracking number %s duplicated") % tracking_number
                    )
                unique_tracking_numbers[default_code].add(tracking_number)
            product_serial.setdefault(default_code, {})[
                tracking_number
            ] = self._prepare_additional_tracking_values(row[2:], company)
        return product_serial

    def import_serials(self):  # noqa: C901
        self.ensure_one()

        msglog = MsgLog(
            {
                "MoreSerialsThanLines": _(
                    "Not all serial numbers have been assigned. "
                    "There are more serial numbers "
                    "in the file than picking lines without them.\n"
                    "Serial numbers not assigned:"
                ),
                "MoreLinesThanSerials": _(
                    "There's not enough serial numbers in input file so there's "
                    "picking lines with operations without serial number:"
                ),
            }
        )

        active_ids = self.env.context.get("active_ids")
        active_model = self.env.context.get("active_model")
        picking = self.env[active_model].browse(active_ids)

        _fname, ext = self.datas_fname.rsplit(".", 1)
        file = base64.b64decode(self.datas)
        if ext == "xls":
            book = xlrd.open_workbook(file_contents=file)
            sheet = book.sheet_by_index(0)
        elif ext == "xlsx":
            book = load_workbook(filename=BytesIO(file))
            sheet = book.worksheets[0]
        else:
            raise UserError(_("Extension %s no supported") % ext)

        sheet_data = getattr(self, "_get_product_serial_%s" % ext)(sheet)
        if not sheet_data:
            raise UserError(_("There's no data in input file"))

        product_serial = self._prepare_data(sheet_data, picking.company_id)

        # set serials on move lines
        for default_code, tracking_values in product_serial.items():
            product = self.env["product.product"].search(
                [
                    ("default_code", "=", default_code),
                    "|",
                    ("company_id", "=", picking.company_id.id),
                    ("company_id", "=", False),
                ]
            )
            if not product:
                raise UserError(_("The product %s does not exist") % default_code)
            if len(product) > 1:
                raise UserError(
                    _("There's more than one product with reference %s") % default_code
                )
            if product.tracking != "serial":
                raise UserError(
                    _("The product %s has no serial tracking type") % default_code
                )

            tracking_numbers = list(tracking_values.keys())

            # get lines of the picking lines
            product_move_lines = picking.move_lines.filtered(
                lambda x: x.product_id.id == product.id
            )
            if not product_move_lines:
                raise UserError(_("There's no lines with product %s") % default_code)

            detail_move_lines = product_move_lines.mapped("move_line_ids")
            if not detail_move_lines:
                raise UserError(
                    _("The line with the product %s has no detail movements created")
                    % default_code
                )

            # classify already imported and pending
            dmls_pending, tns_imported = self.env["stock.move.line"], []
            for dml in detail_move_lines.sorted(lambda x: (x.move_id.sequence, x.id)):
                if dml.product_uom_qty != 1:
                    raise UserError(
                        _("The line with the product %s must have quantity 1")
                        % default_code
                    )
                if not dml.lot_id:
                    dmls_pending += dml
                else:
                    tn = dml.lot_id.name
                    if tn in tracking_numbers:
                        tns_imported.append(tn)

            tns_pending = []
            for tn in tracking_numbers:
                if tn not in tns_imported:
                    tns_pending.append(tn)

            # import
            dmls_n, tns_n = len(dmls_pending), len(tns_pending)
            if not tns_pending and dmls_pending:
                msglog.add_msg(
                    "MoreLinesThanSerials",
                    _("%s -> %i operations empty") % (default_code, dmls_n),
                )
            else:
                tracking_numbers_paired, tracking_numbers_rest = (
                    tns_pending[:dmls_n],
                    tns_pending[dmls_n:],
                )
                for tracking_number, line in zip(tracking_numbers_paired, dmls_pending):
                    # qty_done is always 1 on serial
                    line.qty_done = 1

                    # find/create lot
                    lot_id = self.env["stock.production.lot"].search(
                        [
                            ("company_id", "=", picking.company_id.id),
                            ("product_id", "=", line.product_id.id),
                            ("name", "=", tracking_number),
                        ]
                    )
                    values = tracking_values[tracking_number]
                    if not lot_id:
                        if not picking.picking_type_id.use_create_lots:
                            raise UserError(
                                _(
                                    "The creation of Lot/Serial number is "
                                    "not allowed in this operation type"
                                )
                            )
                        values.update(
                            {
                                "company_id": picking.company_id.id,
                                "product_id": line.product_id.id,
                                "name": tracking_number,
                            }
                        )
                        lot_id = self.env["stock.production.lot"].create(values)
                    else:
                        if not picking.picking_type_id.use_existing_lots:
                            raise UserError(
                                _(
                                    "The use of existing Lot/Serial number is "
                                    "not allowed in this operation type"
                                )
                            )
                        if values:
                            lot_id.write(values)

                    line.lot_id = lot_id
                    line.lot_name = lot_id.name

                if tns_n > dmls_n:
                    msglog.add_msg(
                        "MoreSerialsThanLines",
                        "%s -> %s" % (default_code, tracking_numbers_rest),
                    )
                elif dmls_n > tns_n:
                    msglog.add_msg(
                        "MoreLinesThanSerials",
                        _("%s -> %i operations empty") % (default_code, dmls_n - tns_n),
                    )

        if not msglog.has_msgs:
            self.result = _("Finished successfully")
        else:
            res_blocks = []
            for mtype, msgs in msglog.get_msgs().items():
                res_blocks.append(
                    "* %s\n%s"
                    % (
                        msglog.get_type_desc(mtype),
                        "\n".join(["    > %s" % x for x in msgs]),
                    )
                )
            self.result = "%s\n\n%s" % (
                _("Finished successfully with warnings"),
                "\n\n".join(res_blocks),
            )

        self.state = "result"

        return {}
